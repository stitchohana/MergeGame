#!/usr/bin/env python3
"""Convert all JSON config files to separate .xlsx files under config/xlsx/

Each JSON becomes its own .xlsx (with multiple sheets if nested tables).
Usage: cd config && python json_to_xlsx.py
Output: config/xlsx/*.xlsx
"""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from facility_reward_validation import validate_facility_reward_order
from home_meridian_progression import validate_home_progression
from recipe_validation import validate_no_launcher_ingredients

BASE = Path(__file__).parent
JSON_DIR = BASE / "json_output"
OUT = BASE / "xlsx"
OUT.mkdir(exist_ok=True)

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")


def join_list(lst, sep=";"):
    return sep.join(str(x) for x in lst) if lst else ""


def join_dict_list(lst, sep=";"):
    if not lst:
        return ""
    parts = []
    for d in lst:
        if isinstance(d, dict):
            vals = list(d.values())
            if len(vals) == 2:
                parts.append(f"{vals[0]}:{vals[1]}")
            else:
                parts.append(":".join(str(v) for v in vals))
        else:
            parts.append(str(d))
    return sep.join(parts)


def item_type(item):
    """Read an item type consistently when JSON stores it as int or text."""
    value = item.get("type", 0)
    try:
        return int(value)
    except (TypeError, ValueError):
        return value


def fixed_orders_for_xlsx(fixed_orders):
    """Serialize flat or wave-grouped fixed orders while keeping waves readable."""
    if not fixed_orders:
        return ""
    if (
        isinstance(fixed_orders, list)
        and all(
            isinstance(wave, dict)
            and set(wave.keys()) == {"item_ids"}
            and isinstance(wave.get("item_ids"), list)
            and wave["item_ids"]
            for wave in fixed_orders
        )
    ):
        return "\n".join(
            f"item_ids:{{{','.join(str(item_id) for item_id in wave['item_ids'])}}}"
            for wave in fixed_orders
        )
    if isinstance(fixed_orders, list) and fixed_orders and all(isinstance(wave, list) for wave in fixed_orders):
        wave_text = [json.dumps(wave, ensure_ascii=False, separators=(",", ":")) for wave in fixed_orders]
        return "[\n  " + ",\n  ".join(wave_text) + "\n]"
    return json.dumps(fixed_orders, ensure_ascii=False, separators=(",", ":"))


def format_rewards(obj):
    """Convert a rewards dict like {tokens: [{token:4,amount:2}]} to string '4:2;3:15'"""
    if not obj:
        return ""
    tokens = obj.get("tokens", [])
    items = obj.get("items", [])
    parts = []
    for t in tokens:
        parts.append(f"{t.get('token','?')}:{t.get('amount','?')}")
    for it in items:
        parts.append(f"item:{it.get('id','?')}:{it.get('count','?')}")
    return ";".join(parts)


def reward_json_for_xlsx(obj):
    if isinstance(obj, (dict, list)):
        return json.dumps(obj, ensure_ascii=False, separators=(",", ":"))
    return ""


def add_sheet(ws, headers, rows):
    """Write headers + rows into a worksheet with formatting."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = THIN
            c.alignment = LEFT if ci > 2 else CENTER
    for ci in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16
    ws.freeze_panes = "A2"


def new_book():
    wb = Workbook()
    wb.remove(wb.active)
    return wb


def save(wb, name):
    path = OUT / f"{name}.xlsx"
    try:
        wb.save(path)
        print(f"  -> xlsx/{name}.xlsx ({len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)})")
    except PermissionError:
        print(f"  SKIP xlsx/{name}.xlsx — 文件被占用，请关闭后重试")


# ════════════════════════════════════════════════════════════
#  items.xlsx
# ════════════════════════════════════════════════════════════
print("Building items.xlsx...")
data = json.load(open(JSON_DIR / "items.json", encoding="utf-8"))
wb = new_book()

regular = [it for it in data["regular"] if item_type(it) == 0]
consumables = [it for it in data["regular"] if item_type(it) == 4]

hr = ["id","level","name","icon","group_id","type","describe","value","sell_price"]
rr = []
for it in regular:
    rr.append([it.get("id",""), it.get("level",""), it.get("name",""), it.get("icon",""),
               it.get("group_id",""), it.get("type",""), it.get("describe",""), it.get("value",""),
               it.get("sell_price","")])
ws = wb.create_sheet("items_regular")
add_sheet(ws, hr, rr)

hcons = ["id","level","name","icon","type","describe","value","max_charges","recharge_time","no_cost","spawns(id:weight)","fixed_spawns"]
rcons = []
for it in consumables:
    rcons.append([it.get("id",""), it.get("level",""), it.get("name",""), it.get("icon",""),
                  it.get("type",""), it.get("describe",""), it.get("value",""),
                  it.get("max_charges",""), it.get("recharge_time",""),
                  "TRUE" if it.get("no_cost") else "",
                  join_dict_list(it.get("spawns", [])), join_list(it.get("fixed_spawns", []))])
ws4 = wb.create_sheet("items_recipe_product")
add_sheet(ws4, hcons, rcons)

he = ["id","level","name","icon","group_id","type","describe","effect_type","effect_value"]
re = []
effect_config_items = data.get("effect", []) + [
    it for it in consumables if it.get("effect_type") is not None
]
for it in effect_config_items:
    re.append([it.get("id",""), it.get("level",""), it.get("name",""), it.get("icon",""),
               it.get("group_id",""), it.get("type",""), it.get("describe",""), it.get("effect_type",""), it.get("effect_value","")])
ws5 = wb.create_sheet("items_effect")
add_sheet(ws5, he, re)

hl = ["id","level","name","icon","group_id","type","max_charges","recharge_time","describe","no_cost","spawns(id:weight)","fixed_spawns"]
rl = []
for it in data["launcher"]:
    rl.append([it.get("id",""), it.get("level",""), it.get("name",""), it.get("icon",""),
               it.get("group_id",""), it.get("type",""), it.get("max_charges",""),
               it.get("recharge_time",""), it.get("describe",""),
               "TRUE" if it.get("no_cost") else "",
               join_dict_list(it.get("spawns", [])), join_list(it.get("fixed_spawns", []))])
ws2 = wb.create_sheet("items_launcher")
add_sheet(ws2, hl, rl)

hc = ["id","level","name","group_id","type","icon","describe","recipes"]
rc = []
for it in data["crafting"]:
    rc.append([it.get("id",""), it.get("level",""), it.get("name",""), it.get("group_id",""),
               it.get("type",""), it.get("icon",""), it.get("describe",""), join_list(it.get("recipes", []))])
ws3 = wb.create_sheet("items_crafting")
add_sheet(ws3, hc, rc)

save(wb, "items")

# ════════════════════════════════════════════════════════════
#  meridians.xlsx
# ════════════════════════════════════════════════════════════
print("Building meridians.xlsx...")
data = json.load(open(JSON_DIR / "meridians.json", encoding="utf-8"))
wb = new_book()
h = ["stage","count_min","count_max","acupoint_rewards","order_count","fixed_orders"]
r = []
for t in data["thresholds"]:
    r.append([t.get("stage",""), t.get("count_min",""), t.get("count_max",""),
              t.get("acupoint_rewards",""), t.get("order_count",""),
              fixed_orders_for_xlsx(t.get("fixed_orders", []))])
ws = wb.create_sheet("meridians")
add_sheet(ws, h, r)
for row_index, threshold in enumerate(data["thresholds"], 2):
    if threshold.get("fixed_orders"):
        ws.column_dimensions["F"].width = 80
        ws.cell(row=row_index, column=6).alignment = Alignment(
            horizontal="left", vertical="top", wrap_text=True
        )
        ws.row_dimensions[row_index].height = 100
level_range_headers = [
    "cultivation_min", "cultivation_max",
    "items_regular_min", "items_regular_max",
    "items_recipe_product_min", "items_recipe_product_max",
]
level_range_rows = []
for level_range in data.get("order_level_ranges", []):
    regular_range = level_range.get("items_regular", ["", ""])
    recipe_product_range = level_range.get("items_recipe_product", ["", ""])
    level_range_rows.append([
        level_range.get("cultivation_min", ""), level_range.get("cultivation_max", ""),
        regular_range[0], regular_range[1],
        recipe_product_range[0], recipe_product_range[1],
    ])
ws2 = wb.create_sheet("order_level_ranges")
add_sheet(ws2, level_range_headers, level_range_rows)
save(wb, "meridians")

# ════════════════════════════════════════════════════════════
#  rewards.xlsx
# ════════════════════════════════════════════════════════════
print("Building rewards.xlsx...")
data = json.load(open(JSON_DIR / "rewards.json", encoding="utf-8"))
wb = new_book()
h = ["reward_id","tokens(token:amount)","items(id:count)"]
r = []
for rid, rd in data["rewards"].items():
    r.append([int(rid), join_dict_list(rd.get("tokens", [])), join_dict_list(rd.get("items", []))])
ws = wb.create_sheet("rewards")
add_sheet(ws, h, r)
save(wb, "rewards")

# ════════════════════════════════════════════════════════════
#  game_config.xlsx
# ════════════════════════════════════════════════════════════
print("Building game_config.xlsx...")
data = json.load(open(JSON_DIR / "game_config.json", encoding="utf-8"))
def flatten(obj, prefix=""):
    result = []
    for k, v in obj.items():
        full = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            result.extend(flatten(v, full))
        else:
            result.append([full, v])
    return result
wb = new_book()
ws = wb.create_sheet("game_config")
add_sheet(ws, ["parameter","value"], flatten(data))
save(wb, "game_config")

# ════════════════════════════════════════════════════════════
#  recipes.xlsx
# ════════════════════════════════════════════════════════════
print("Building recipes.xlsx...")
data = json.load(open(JSON_DIR / "recipes.json", encoding="utf-8"))
items_data = json.load(open(JSON_DIR / "items.json", encoding="utf-8"))
validate_no_launcher_ingredients(data["recipes"], items_data.get("launcher", []))
wb = new_book()
h = ["id","name","ingredients","result","craft_time"]
r = []
for rc in data["recipes"]:
    r.append([rc.get("id",""), rc.get("name",""), join_list(rc.get("ingredients",[])),
              rc.get("result",""), rc.get("craft_time","")])
ws = wb.create_sheet("recipes")
add_sheet(ws, h, r)
save(wb, "recipes")

# ════════════════════════════════════════════════════════════
#  initial_setup.xlsx
# ════════════════════════════════════════════════════════════
print("Building initial_setup.xlsx...")
data = json.load(open(JSON_DIR / "initial_setup.json", encoding="utf-8"))
wb = new_book()
h = ["section","id","col","row","immovable"]
r = []
for section, sec in data.items():
    for it in sec.get("items", []):
        r.append([section, it.get("id",""), it.get("col",""), it.get("row",""),
                  "TRUE" if it.get("immovable") else "FALSE"])
ws = wb.create_sheet("initial_setup")
add_sheet(ws, h, r)
save(wb, "initial_setup")

# ════════════════════════════════════════════════════════════
#  cultivation.xlsx
# ════════════════════════════════════════════════════════════
print("Building cultivation.xlsx...")
data = json.load(open(JSON_DIR / "cultivation.json", encoding="utf-8"))
wb = new_book()
h = ["stage_index","name","exp","max_qi","breakthrough_items","breakthrough_reward_id"]
r = []
for i, s in enumerate(data["stages"]):
    r.append([i, s.get("name",""), s.get("exp",""), s.get("max_qi",""),
              join_dict_list(s.get("breakthrough_items", [])), s.get("breakthrough_reward_id","")])
ws = wb.create_sheet("cultivation")
add_sheet(ws, h, r)
save(wb, "cultivation")

# ════════════════════════════════════════════════════════════
#  shop.xlsx
# ════════════════════════════════════════════════════════════
print("Building shop.xlsx...")
data = json.load(open(JSON_DIR / "shop.json", encoding="utf-8"))
wb = new_book()
ws = wb.create_sheet("shop")
add_sheet(ws, ["id","price"], [[it["id"], it["price"]] for it in data["items"]])
save(wb, "shop")

# ════════════════════════════════════════════════════════════
#  expedition.xlsx
# ════════════════════════════════════════════════════════════
print("Building expedition.xlsx...")
data = json.load(open(JSON_DIR / "expedition.json", encoding="utf-8"))
wb = new_book()
h = ["id","name","hp","atk","accept_effect_ids","loot","describe"]
r = []
for m in data["monsters"]:
    r.append([m.get("id",""), m.get("name",""), m.get("hp",""), m.get("atk",""),
              join_list(m.get("accept_effect_ids",[])), join_list(m.get("loot",[])),
              m.get("describe","")])
ws = wb.create_sheet("monsters")
add_sheet(ws, h, r)

ws2 = wb.create_sheet("maps")
add_sheet(ws2, ["id","name","describe"],
    [[mp.get("id",""), mp.get("name",""), mp.get("describe","")] for mp in data["maps"]])

h3 = ["map_id","stage","name","monsters(monster_id:count)","boss_monster_id"]
r3 = []
for mp in data["maps"]:
    for st in mp.get("stages", []):
        boss = st.get("boss", {})
        r3.append([mp["id"], st.get("stage",""), st.get("name",""),
                   join_dict_list(st.get("monsters",[])),
                   boss.get("monster_id","") if boss else ""])
ws3 = wb.create_sheet("map_stages")
add_sheet(ws3, h3, r3)
save(wb, "expedition")


# ════════════════════════════════════════════════════════════
#  tokens.xlsx
# ════════════════════════════════════════════════════════════
print("Building tokens.xlsx...")
data = json.load(open(JSON_DIR / "tokens.json", encoding="utf-8"))
wb = new_book()
ws = wb.create_sheet("tokens")
add_sheet(ws, ["id","name","icon","describe"],
    [[t.get("id",""), t.get("name",""), t.get("icon",""), t.get("describe","")]
     for t in data["tokens"]])
save(wb, "tokens")

# ════════════════════════════════════════════════════════════
#  server.xlsx
# ════════════════════════════════════════════════════════════
print("Building server.xlsx...")
data = json.load(open(JSON_DIR / "server.json", encoding="utf-8"))
wb = new_book()
ws = wb.create_sheet("server")
add_sheet(ws, ["parameter","value"], [[k, v] for k, v in data.items()])
save(wb, "server")

# ════════════════════════════════════════════════════════════
#  quests.xlsx
# ════════════════════════════════════════════════════════════
print("Building quests.xlsx...")
data = json.load(open(JSON_DIR / "quests.json", encoding="utf-8"))
wb = new_book()
h = ["id","type","name","description","target_count","rewards","auto_reward","reset_cycle"]
r = []
for q in data["quests"]:
    r.append([q.get("id",""), q.get("type",""), q.get("name",""), q.get("description",""),
              q.get("target_count",""), q.get("rewards",""),
              "TRUE" if q.get("auto_reward") else "FALSE", q.get("reset_cycle","")])
ws = wb.create_sheet("quests")
add_sheet(ws, h, r)
save(wb, "quests")

# ════════════════════════════════════════════════════════════
#  activities.xlsx
# ════════════════════════════════════════════════════════════
print("Building activities.xlsx...")
data = json.load(open(JSON_DIR / "activities.json", encoding="utf-8"))
wb = new_book()
h = ["id","name","cycle","start_time","end_time","widget"]
r = []
for a in data["activities"]:
    r.append([a.get("id",""), a.get("name",""), a.get("cycle",""),
              a.get("start_time",""), a.get("end_time",""), a.get("widget","")])
ws = wb.create_sheet("activities")
add_sheet(ws, h, r)
save(wb, "activities")

# ════════════════════════════════════════════════════════════
#  weekly_tasks.xlsx
# ════════════════════════════════════════════════════════════
print("Building weekly_tasks.xlsx...")
data = json.load(open(JSON_DIR / "weekly_tasks.json", encoding="utf-8"))
wb = new_book()
h = ["activity_id","day","quests"]
r = []
for wt in data["weekly_tasks"]:
    for day_idx, quests in enumerate(wt.get("daily_quests", []), 1):
        r.append([wt.get("activity_id",""), day_idx, join_list(quests)])
ws = wb.create_sheet("weekly_tasks")
add_sheet(ws, h, r)
save(wb, "weekly_tasks")

# ════════════════════════════════════════════════════════════
#  home_meridians.xlsx
# ════════════════════════════════════════════════════════════
print("Building home_meridians.xlsx...")
data = json.load(open(JSON_DIR / "home_meridians.json", encoding="utf-8"))
items_for_facility_validation = json.load(open(JSON_DIR / "items.json", encoding="utf-8"))
recipes_for_facility_validation = json.load(open(JSON_DIR / "recipes.json", encoding="utf-8"))["recipes"]
setup_for_facility_validation = json.load(open(JSON_DIR / "initial_setup.json", encoding="utf-8"))
rewards_for_facility_validation = json.load(open(JSON_DIR / "rewards.json", encoding="utf-8")).get("rewards", {})
cultivation_for_home_validation = json.load(open(JSON_DIR / "cultivation.json", encoding="utf-8"))["stages"]
validate_home_progression(
    data["stages"],
    cultivation_for_home_validation,
    rewards_for_facility_validation,
    items_for_facility_validation,
)
validate_facility_reward_order(
    data["stages"],
    items_for_facility_validation,
    recipes_for_facility_validation,
    setup_for_facility_validation,
    rewards_for_facility_validation,
    cultivation_for_home_validation,
)
wb = new_book()
ws = wb.create_sheet("home_meridians")
add_sheet(ws, ["cultivation_level","name","acupoints","qi_cost","acupoint_exp","circulation_reward"],
    [[s.get("cultivation_level",""), s.get("name",""), s.get("acupoints",""), s.get("qi_cost",""),
      s.get("acupoint_exp", 0),
      reward_json_for_xlsx(s.get("circulation_reward", {}))]
     for s in data["stages"]])
save(wb, "home_meridians")

print(f"\nDone! {len(list(OUT.glob('*.xlsx')))} files in {OUT}/")
