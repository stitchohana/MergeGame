#!/usr/bin/env python3
"""Convert separate .xlsx files under config/xlsx/ back to JSON config files.

Usage: cd config && python xlsx_to_json.py
Output: config/json_output/
"""

import json, re
from pathlib import Path
from openpyxl import load_workbook

from facility_reward_validation import validate_facility_reward_order
from home_meridian_progression import SPIRIT_STONE_MINE_ID, validate_home_progression
from recipe_validation import validate_no_launcher_ingredients

BASE = Path(__file__).parent
XLSX_DIR = BASE / "xlsx"
OUT = BASE / "json_output"
OUT.mkdir(exist_ok=True)

ITEM_ICON_DIRECTORY = "icons_simple"
SPIRIT_STONE_MINE_FAMILY = SPIRIT_STONE_MINE_ID // 100

CRAFT_TIME_SECONDS_BY_PRODUCT_LEVEL = {
    1: 60,
    2: 180,
    3: 300,
    4: 600,
    5: 1200,
    6: 1800,
    7: 2700,
    8: 3600,
    9: 5400,
    10: 7200,
    11: 10800,
    12: 14400,
    13: 21600,
    14: 28800,
    15: 36000,
    16: 43200,
}


def item_icon_path(item_id):
    return f"res://assets/items/{ITEM_ICON_DIRECTORY}/{item_id}.png"


def normalized_item_icon(row):
    """Keep an explicit workbook icon; fill the canonical path only when blank."""
    icon = str(row.get("icon", "")).strip()
    if icon:
        return icon
    item_id = parse_int(row.get("id"))
    return item_icon_path(item_id) if item_id is not None else ""


def parse_int(v):
    if v is None or v == "" or v == "None":
        return None
    return int(float(v))


def parse_bool(v):
    if v is None or v == "":
        return False
    if isinstance(v, bool):
        return v
    return str(v).upper() in ("TRUE", "1", "YES")


def parse_int_list(v, sep=";"):
    if v is None or v == "":
        return []
    return [int(x.strip()) for x in str(v).split(sep) if x.strip()]

def parse_ingredient_list(v):
    if v is None or v == "":
        return []
    import re
    result = []
    for part in str(v).split(";"):
        part = part.strip()
        if not part:
            continue
        # Handle dict-like: {'id': 1002, 'count': 3} or JSON: {"id": 1002, "count": 3}
        if part.startswith("{"):
            m = re.search(r"['\"]id['\"]\s*:\s*(\d+)", part)
            if m:
                result.append(int(m.group(1)))
            continue
        # Handle id:count format: 1002:3
        if ":" in part:
            iid, _ = part.split(":", 1)
            result.append(int(iid.strip()))
            continue
        result.append(int(part))
    return result


def parse_dict_list(v, key1="id", key2="weight", sep=";"):
    if v is None or v == "":
        return []
    result = []
    for part in str(v).split(sep):
        part = part.strip()
        if not part:
            continue
        if ":" in part:
            a, b = part.split(":", 1)
            result.append({key1: int(a), key2: int(b)})
        else:
            result.append(int(part))
    return result


def parse_reward_config(v):
    if v is None or v == "":
        return None
    text = str(v).strip()
    if text.startswith("{") or text.startswith("["):
        return json.loads(text)
    if re.fullmatch(r"\d+", text):
        return parse_int(text)
    return None


def parse_fixed_orders(v):
    if v is None or v == "":
        return []
    text = str(v).strip()

    # XLSX shorthand: one line per wave, for example
    # item_ids:{5003,5004,6001}
    shorthand_lines = [line.strip() for line in text.splitlines() if line.strip()]
    if shorthand_lines and all(line.lower().startswith("item_ids") for line in shorthand_lines):
        waves = []
        for line in shorthand_lines:
            match = re.fullmatch(r"item_ids\s*:\s*\{([^{}]*)\}", line, re.IGNORECASE)
            if not match:
                break
            ids = [parse_int(part.strip()) for part in re.split(r"[,;]", match.group(1)) if part.strip()]
            if any(item_id is None for item_id in ids):
                raise ValueError(f"invalid fixed_orders shorthand: {line}")
            if ids:
                waves.append({"item_ids": ids})
        else:
            return waves

    if text.startswith("["):
        parsed = json.loads(text)
        if not isinstance(parsed, list):
            return []
        def parse_order(entry):
            if isinstance(entry, dict):
                if "item_id" in entry and "item_ids" not in entry:
                    item_id = parse_int(entry.get("item_id"))
                    if item_id is None:
                        return None
                    order = dict(entry)
                    order["item_id"] = item_id
                    return order
                raw_ids = entry.get("item_ids", [])
                item_ids = parse_int_list(raw_ids) if isinstance(raw_ids, str) else [int(x) for x in raw_ids]
                order = {key: value for key, value in entry.items() if key != "item_ids"}
                order["item_ids"] = item_ids
                return order if item_ids else None
            if isinstance(entry, list):
                item_ids = [int(x) for x in entry]
                return {"item_ids": item_ids} if item_ids else None
            item_id = int(entry)
            return {"item_ids": [item_id]}

        # A nested list is the wave-aware form:
        # [[{"item_id": 5003}], [{"item_id": 5004}]]
        # Keep the old flat list form (including [[5003], [5004]]) valid and
        # treat it as one wave.
        is_wave_config = any(
            isinstance(entry, list)
            and any(
                isinstance(order, dict)
                or (isinstance(order, list) and any(isinstance(value, (dict, list)) for value in order))
                for order in entry
            )
            for entry in parsed
        )
        if is_wave_config:
            waves = []
            for wave in parsed:
                parsed_wave = [order for entry in wave if (order := parse_order(entry))]
                if parsed_wave:
                    waves.append(parsed_wave)
            return waves

        result = [order for entry in parsed if (order := parse_order(entry))]
        return result
    return [{"item_ids": [int(x.strip())]} for x in text.split(";") if x.strip()]


def read_rows(ws):
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                d[headers[i]] = str(val).strip() if val is not None and str(val).strip() != "None" else ""
        if any(v != "" for v in d.values()):
            rows.append(d)
    return rows


def save_json(filename, data):
    path = OUT / filename
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  -> {filename}")


def open_book(name):
    return load_workbook(XLSX_DIR / f"{name}.xlsx")


def read_recipe_definitions():
    recipe_wb = open_book("recipes")
    recipes = []
    for row in read_rows(recipe_wb["recipes"]):
        recipes.append({
            "id": parse_int(row["id"]), "name": row["name"],
            "ingredients": parse_ingredient_list(row["ingredients"]),
            "result": parse_int(row["result"]),
            "craft_time": parse_int(row["craft_time"]),
        })
    return recipes


def derive_recipe_product_levels(base_levels, product_ids, recipes):
    """Return each recipe product's max required material level."""
    recipes_by_result = {}
    for recipe in recipes:
        recipes_by_result.setdefault(recipe["result"], []).append(recipe)

    cache = {}

    def resolve(item_id, visiting=None):
        if item_id in cache:
            return cache[item_id]
        if item_id in base_levels and item_id not in product_ids:
            cache[item_id] = base_levels[item_id]
            return cache[item_id]
        if item_id not in product_ids:
            # The recipe source uses 220xx as virtual references to a pill
            # tier even though those intermediate items are not item rows.
            if 22000 < item_id < 22100:
                cache[item_id] = item_id - 22000
                return cache[item_id]
            return None

        if visiting is None:
            visiting = set()
        if item_id in visiting:
            return None
        visiting.add(item_id)
        ingredient_levels = []
        for recipe in recipes_by_result.get(item_id, []):
            ingredient_levels.extend(
                resolve(ingredient_id, visiting)
                for ingredient_id in recipe.get("ingredients", [])
            )
        visiting.remove(item_id)

        level = max(ingredient_levels) if ingredient_levels and all(
            value is not None for value in ingredient_levels
        ) else None
        cache[item_id] = level
        return level

    return {item_id: resolve(item_id) for item_id in product_ids}


# ─── items ─────────────────────────────────────────────────
print("Building items.json...")
wb = open_book("items")
recipe_definitions = read_recipe_definitions()

regular = []
for row in read_rows(wb["items_regular"]):
    item = {"id": parse_int(row["id"]), "level": parse_int(row["level"]),
            "name": row["name"], "icon": normalized_item_icon(row),
            "group_id": parse_int(row["group_id"]), "describe": row["describe"]}
    if row.get("type") != "":
        item["type"] = parse_int(row["type"])
    if row.get("value"):
        item["value"] = parse_int(row["value"])
    if row.get("sell_price"):
        item["sell_price"] = parse_int(row["sell_price"])

    regular.append(item)

launcher = []
for row in read_rows(wb["items_launcher"]):
    item = {"id": parse_int(row["id"]), "level": parse_int(row["level"]),
            "name": row["name"], "icon": normalized_item_icon(row),
            "group_id": parse_int(row["group_id"]),
            "max_charges": parse_int(row["max_charges"]),
            "recharge_time": parse_int(row["recharge_time"]),
            "describe": row["describe"]}
    if row.get("type"):
        item["type"] = row["type"]
    if parse_bool(row.get("no_cost")):
        item["no_cost"] = True
    spawns = parse_dict_list(row.get("spawns(id:weight)", ""))
    fixed = parse_int_list(row.get("fixed_spawns", ""))
    if fixed:
        item["fixed_spawns"] = fixed
    elif spawns:
        item["spawns"] = spawns
    launcher.append(item)

validate_no_launcher_ingredients(recipe_definitions, launcher)

crafting = []
for row in read_rows(wb["items_crafting"]):
    item = {
        "id": parse_int(row["id"]), "level": parse_int(row["level"]),
        "name": row["name"], "group_id": parse_int(row["group_id"]),
        "icon": normalized_item_icon(row), "describe": row["describe"],
        "recipes": parse_int_list(row.get("recipes", "")),
    }
    if row.get("type") != "":
        item["type"] = parse_int(row["type"])
    crafting.append(item)
# Read items_recipe_product sheet if it exists
if "items_recipe_product" in [ws.title for ws in wb.worksheets]:
    recipe_product_rows = read_rows(wb["items_recipe_product"])
    recipe_product_ids = {parse_int(row["id"]) for row in recipe_product_rows}
    base_levels = {}
    for item in [*regular, *launcher, *crafting]:
        if item["id"] not in recipe_product_ids and item.get("level") is not None:
            base_levels[item["id"]] = item["level"]
    recipe_product_levels = derive_recipe_product_levels(
        base_levels, recipe_product_ids, recipe_definitions
    )
    for row in recipe_product_rows:
        item_id = parse_int(row["id"])
        # The workbook is the source of truth.  Derivation remains a fallback
        # for legacy sheets that predate the explicit level column values.
        recipe_level = parse_int(row.get("level"))
        if recipe_level is None:
            recipe_level = recipe_product_levels.get(item_id)
        item = {"id": item_id,
                "level": recipe_level,
                "name": row["name"], "icon": normalized_item_icon(row),
                "describe": row["describe"],
                "type": 4}
        if row.get("value"):
            item["value"] = parse_int(row["value"])
        # Optional launcher fields
        if row.get("max_charges"):
            item["max_charges"] = parse_int(row["max_charges"])
        if row.get("recharge_time"):
            item["recharge_time"] = parse_int(row["recharge_time"])
        if parse_bool(row.get("no_cost")):
            item["no_cost"] = True
        spawns = parse_dict_list(row.get("spawns(id:weight)", ""))
        fixed = parse_int_list(row.get("fixed_spawns", ""))
        if fixed:
            item["fixed_spawns"] = fixed
        elif spawns:
            item["spawns"] = spawns
        regular.append(item)
effect_items = []
if "items_effect" in [ws.title for ws in wb.worksheets]:
    regular_by_id = {item["id"]: item for item in regular}
    for row in read_rows(wb["items_effect"]):
        item_id = parse_int(row["id"])
        item_type = parse_int(row.get("type")) or 5
        if item_type == 4 and item_id in regular_by_id:
            regular_by_id[item_id]["effect_type"] = parse_int(row["effect_type"])
            regular_by_id[item_id]["effect_value"] = parse_int(row["effect_value"])
            continue
        item = {"id": parse_int(row["id"]), "level": parse_int(row["level"]),
                "name": row["name"], "icon": normalized_item_icon(row),
                "group_id": parse_int(row["group_id"]), "describe": row["describe"],
                "effect_type": parse_int(row["effect_type"]),
                "effect_value": parse_int(row["effect_value"]),
                "type": 5}

        if row.get("value"):
            item["value"] = parse_int(row["value"])

        effect_items.append(item)
save_json("items.json", {"regular": regular, "launcher": launcher, "crafting": crafting, "effect": effect_items})

# ─── meridians ─────────────────────────────────────────────
print("Building meridians.json...")
wb = open_book("meridians")
thresholds = []
for row in read_rows(wb["meridians"]):
    threshold = {
        "stage": parse_int(row["stage"]),
        "count_min": parse_int(row["count_min"]),
        "count_max": parse_int(row["count_max"]),
        "acupoint_rewards": parse_int(row.get("acupoint_rewards", "")),
        "order_count": parse_int(row["order_count"]),
    }
    fixed_orders = parse_fixed_orders(row.get("fixed_orders", ""))
    if fixed_orders:
        threshold["fixed_orders"] = fixed_orders
    thresholds.append(threshold)
order_level_ranges = []
for row in read_rows(wb["order_level_ranges"]):
    order_level_ranges.append({
        "cultivation_min": parse_int(row["cultivation_min"]),
        "cultivation_max": parse_int(row["cultivation_max"]),
        "items_regular": [
            parse_int(row["items_regular_min"]),
            parse_int(row["items_regular_max"]),
        ],
        "items_recipe_product": [
            parse_int(row["items_recipe_product_min"]),
            parse_int(row["items_recipe_product_max"]),
        ],
    })
save_json("meridians.json", {
    "order_level_ranges": order_level_ranges,
    "thresholds": thresholds,
})

# ─── rewards ───────────────────────────────────────────────
print("Building rewards.json...")
wb = open_book("rewards")
rewards = {}
for row in read_rows(wb["rewards"]):
    entry = {}
    tokens = parse_dict_list(row.get("tokens(token:amount)", ""), "token", "amount")
    if tokens:
        entry["tokens"] = tokens
    items = parse_dict_list(row.get("items(id:count)", ""), "id", "count")
    if items:
        entry["items"] = items
    if entry:
        rewards[row["reward_id"]] = entry
save_json("rewards.json", {"rewards": rewards})

# ─── game_config ───────────────────────────────────────────
print("Building game_config.json...")
wb = open_book("game_config")
def unflatten(rows):
    result = {}
    for row in rows:
        key = row["parameter"]
        val = row["value"]
        if not key or val == "":
            continue
        v = parse_int(val) if val.lstrip("-").isdigit() else val
        parts = key.split(".")
        target = result
        for part in parts[:-1]:
            if part not in target:
                target[part] = {}
            target = target[part]
        target[parts[-1]] = v
    return result
game_config = unflatten(read_rows(wb["game_config"]))
# Migrate workbooks created before speedup billing switched from seconds to
# minutes. The current design is one spirit stone per started minute.
if "craft_speedup_stone_cost_per_second" in game_config:
    game_config.pop("craft_speedup_stone_cost_per_second")
    game_config.setdefault("craft_speedup_stone_cost_per_minute", 1)
if "launcher_speedup_stone_cost_per_second" in game_config:
    game_config.pop("launcher_speedup_stone_cost_per_second")
    game_config.setdefault("launcher_speedup_stone_cost_per_minute", 1)
save_json("game_config.json", game_config)

# ─── recipes ───────────────────────────────────────────────
print("Building recipes.json...")
recipe_result_levels = {
    item["id"]: item.get("level")
    for item in regular
}
for recipe in recipe_definitions:
    product_level = recipe_result_levels.get(recipe["result"])
    expected_craft_time = CRAFT_TIME_SECONDS_BY_PRODUCT_LEVEL.get(product_level)
    if expected_craft_time is None:
        raise ValueError(
            f"recipe {recipe['id']} result {recipe['result']} has unsupported "
            f"product level {product_level}"
        )
    if recipe["craft_time"] != expected_craft_time:
        raise ValueError(
            f"recipe {recipe['id']} level {product_level} craft time must be "
            f"{expected_craft_time}s, got {recipe['craft_time']}s"
        )
save_json("recipes.json", {"recipes": recipe_definitions})

# ─── initial_setup ─────────────────────────────────────────
print("Building initial_setup.json...")
wb = open_book("initial_setup")
setup = {}
for row in read_rows(wb["initial_setup"]):
    section = row["section"]
    if section not in setup:
        setup[section] = {"items": []}
    setup[section]["items"].append({
        "id": parse_int(row["id"]), "col": parse_int(row["col"]),
        "row": parse_int(row["row"]), "immovable": parse_bool(row["immovable"]),
    })
save_json("initial_setup.json", setup)

# ─── cultivation ───────────────────────────────────────────
print("Building cultivation.json...")
wb = open_book("cultivation")
stages = []
for row in read_rows(wb["cultivation"]):
    stage = {"name": row["name"], "exp": parse_int(row["exp"]),
             "max_qi": parse_int(row["max_qi"])}
    breakthrough_items = parse_dict_list(row.get("breakthrough_items", ""), "item_id", "count")
    if breakthrough_items:
        stage["breakthrough_items"] = [
            {"item_id": int(entry["item_id"]), "count": int(entry["count"])}
            for entry in breakthrough_items
            if isinstance(entry, dict) and int(entry.get("item_id", 0)) > 0 and int(entry.get("count", 0)) > 0
        ]
    if row.get("breakthrough_reward_id"):
        stage["breakthrough_reward_id"] = parse_int(row["breakthrough_reward_id"])
    stages.append(stage)
save_json("cultivation.json", {"stages": stages})

# ─── shop ──────────────────────────────────────────────────
print("Building shop.json...")
wb = open_book("shop")
items_list = []
for row in read_rows(wb["shop"]):
    items_list.append({"id": parse_int(row["id"]), "price": parse_int(row["price"])})
save_json("shop.json", {"items": items_list})

# ─── expedition ────────────────────────────────────────────
print("Building expedition.json...")
wb = open_book("expedition")
monsters = []
for row in read_rows(wb["monsters"]):
    monsters.append({
        "id": parse_int(row["id"]), "name": row["name"],
        "hp": parse_int(row["hp"]), "atk": parse_int(row["atk"]),
        "accept_effect_ids": parse_int_list(row["accept_effect_ids"]),
        "loot": parse_int_list(row["loot"]),
        "describe": row["describe"],
    })

maps_raw = {row["id"]: row for row in read_rows(wb["maps"])}
stages_by_map = {}
for row in read_rows(wb["map_stages"]):
    mid = row["map_id"]
    if mid not in stages_by_map:
        stages_by_map[mid] = []
    stage = {
        "stage": parse_int(row["stage"]), "name": row["name"],
        "monsters": parse_dict_list(row.get("monsters(monster_id:count)", ""), "monster_id", "count"),
    }
    boss_id = row.get("boss_monster_id", "")
    if boss_id:
        stage["boss"] = {"monster_id": parse_int(boss_id)}
    stages_by_map[mid].append(stage)

maps_list = []
for mid, mp in maps_raw.items():
    maps_list.append({
        "id": parse_int(mp["id"]), "name": mp["name"],
        "describe": mp["describe"],
        "stages": stages_by_map.get(mid, []),
    })
save_json("expedition.json", {"monsters": monsters, "maps": maps_list})


# ─── tokens ────────────────────────────────────────────────
print("Building tokens.json...")
wb = open_book("tokens")
tokens = []
for row in read_rows(wb["tokens"]):
    tokens.append({
        "id": parse_int(row["id"]), "name": row["name"],
        "icon": row.get("icon", ""), "describe": row["describe"],
    })
save_json("tokens.json", {"tokens": tokens})

# ─── server ────────────────────────────────────────────────
print("Building server.json...")
wb = open_book("server")
server = {}
for row in read_rows(wb["server"]):
    val = row["value"]
    server[row["parameter"]] = parse_int(val) if val.isdigit() else val
save_json("server.json", server)

# ─── quests ────────────────────────────────────────────────
print("Building quests.json...")
wb = open_book("quests")
quests = []
for row in read_rows(wb["quests"]):
    quests.append({
        "id": parse_int(row["id"]), "type": parse_int(row["type"]),
        "name": row["name"], "description": row["description"],
        "target_count": parse_int(row["target_count"]),
        "rewards": parse_int(row["rewards"]),
        "auto_reward": parse_bool(row["auto_reward"]),
        "reset_cycle": parse_int(row["reset_cycle"]),
    })
save_json("quests.json", {"quests": quests})

# ─── activities ────────────────────────────────────────────
print("Building activities.json...")
wb = open_book("activities")
activities = []
for row in read_rows(wb["activities"]):
    act = {"id": parse_int(row["id"]), "name": row["name"],
           "cycle": parse_int(row["cycle"]), "widget": row.get("widget", "")}
    if row.get("start_time"):
        act["start_time"] = row["start_time"]
    if row.get("end_time"):
        act["end_time"] = row["end_time"]
    activities.append(act)
save_json("activities.json", {"activities": activities})

# ─── weekly_tasks ───────────────────────────────────────────
print("Building weekly_tasks.json...")
wb = open_book("weekly_tasks")
tasks_by_activity = {}
for row in read_rows(wb["weekly_tasks"]):
    aid = row["activity_id"]
    if aid not in tasks_by_activity:
        tasks_by_activity[aid] = {"activity_id": parse_int(aid), "daily_quests": []}
    tasks_by_activity[aid]["daily_quests"].append(parse_int_list(row["quests"]))
weekly_tasks = list(tasks_by_activity.values())
save_json("weekly_tasks.json", {"weekly_tasks": weekly_tasks})

# ─── home_meridians ────────────────────────────────────────
print("Building home_meridians.json...")
wb = open_book("home_meridians")
home_stages = []
for row in read_rows(wb["home_meridians"]):
    acupoint_exp = parse_int(row.get("acupoint_exp", "")) or 0
    circulation_reward = parse_reward_config(row.get("circulation_reward", ""))
    if circulation_reward is None:
        raise ValueError(f"home meridian {row['name']} is missing circulation_reward")
    if isinstance(circulation_reward, int) and str(circulation_reward) not in rewards:
        raise ValueError(
            f"home meridian {row['name']} references missing reward ID "
            f"{circulation_reward}"
        )
    home_stages.append({
        "cultivation_level": parse_int(row.get("cultivation_level")),
        "name": row["name"], "acupoints": parse_int(row["acupoints"]),
        "qi_cost": parse_int(row["qi_cost"]),
        "acupoint_exp": acupoint_exp,
        "circulation_reward": circulation_reward,
    })


def reward_exp_amount(reward_config):
    """Return cultivation EXP (token 4) granted by one reward config."""
    if isinstance(reward_config, int):
        reward_config = rewards.get(str(reward_config), {})
    if not isinstance(reward_config, dict):
        return 0
    return sum(
        int(token.get("amount", 0))
        for token in reward_config.get("tokens", [])
        if isinstance(token, dict) and int(token.get("token", 0)) == 4
    )


def home_stage_exp_total(home_stage):
    acupoint_total = home_stage["acupoints"] * int(home_stage.get("acupoint_exp", 0))
    return acupoint_total + reward_exp_amount(home_stage["circulation_reward"])


validate_home_progression(home_stages, stages, rewards, {"regular": regular, "launcher": launcher, "crafting": crafting})


def validate_facility_rewards():
    """Validate the staged facility rewards after the schema migration."""
    validate_facility_reward_order(
        home_stages,
        {"regular": regular, "launcher": launcher, "crafting": crafting},
        recipe_definitions,
        setup,
        rewards,
        stages,
    )
    crafting_by_id = {item["id"]: item for item in crafting}
    recipes_by_id = {recipe["id"]: recipe for recipe in recipe_definitions}
    regular_by_id = {
        item["id"]: item
        for item in regular
        if item.get("group_id") is not None and item.get("level") is not None
    }
    regular_by_group = {}
    for item in regular_by_id.values():
        regular_by_group.setdefault(item["group_id"], []).append(item)

    launcher_outputs = {}
    for item in launcher:
        direct_outputs = set(item.get("fixed_spawns", []))
        direct_outputs.update(spawn["id"] for spawn in item.get("spawns", []))
        reachable_outputs = set(direct_outputs)
        for output_id in direct_outputs:
            output = regular_by_id.get(output_id)
            if output is None:
                continue
            reachable_outputs.update(
                candidate["id"]
                for candidate in regular_by_group.get(output["group_id"], [])
                if candidate["level"] >= output["level"]
            )
        launcher_outputs[item["id"]] = reachable_outputs

    def table_is_usable(item_id, available_materials):
        table = crafting_by_id.get(item_id)
        if table is None:
            return True
        return any(
            set(recipes_by_id[recipe_id]["ingredients"]).issubset(available_materials)
            for recipe_id in table.get("recipes", [])
            if recipe_id in recipes_by_id
        )

    def launcher_has_usable_table(item_id, available_materials, available_tables):
        launcher_materials = launcher_outputs.get(item_id, set())
        if not launcher_materials:
            return True
        for table_id in available_tables:
            table = crafting_by_id.get(table_id)
            if table is None:
                continue
            for recipe_id in table.get("recipes", []):
                recipe = recipes_by_id.get(recipe_id)
                if recipe is None:
                    continue
                ingredients = set(recipe["ingredients"])
                if ingredients.issubset(available_materials) and ingredients.intersection(launcher_materials):
                    return True
        return False

    def iter_reward_items(reward_config):
        if isinstance(reward_config, int):
            reward_config = rewards.get(str(reward_config), {})
        if isinstance(reward_config, dict):
            configs = [reward_config]
        elif isinstance(reward_config, list):
            configs = [entry for entry in reward_config if isinstance(entry, dict)]
        else:
            configs = []
        for config in configs:
            for item in config.get("items", []):
                if isinstance(item, dict) and "id" in item:
                    yield item

    available_materials = set()
    available_launcher_ids = set()
    available_table_ids = set()

    def add_available_facility(item_id):
        if item_id in launcher_outputs:
            available_launcher_ids.add(item_id)
            available_materials.update(launcher_outputs[item_id])
        if item_id in crafting_by_id:
            available_table_ids.add(item_id)

    setup_ids = set()
    for item in setup.get("main", {}).get("items", []):
        item_id = int(item["id"])
        setup_ids.add(item_id)
        add_available_facility(item_id)

    starter_facilities = {12001, 16001, 17001}
    missing_starters = sorted(starter_facilities - setup_ids)
    if missing_starters:
        raise ValueError(f"initial setup is missing tutorial facilities: {missing_starters}")

    mine_circulation_ids = []
    for stage_index, home_stage in enumerate(home_stages):
        seen_ids = set()
        stage_items = list(iter_reward_items(home_stage.get("circulation_reward")))
        for item in stage_items:
            item_id = parse_int(item.get("id"))
            count = parse_int(item.get("count", 1)) or 0
            if item_id is None or count <= 0:
                raise ValueError(f"home meridian {stage_index} has an invalid facility reward")
            if item_id in seen_ids:
                raise ValueError(f"home meridian {stage_index} repeats facility {item_id}")
            seen_ids.add(item_id)
            if item_id not in launcher_outputs and item_id not in crafting_by_id:
                raise ValueError(f"home meridian {stage_index} references unknown facility {item_id}")
            add_available_facility(item_id)

        for item in stage_items:
            item_id = int(item["id"])
            count = parse_int(item.get("count", 1)) or 0
            if not table_is_usable(item_id, available_materials):
                raise ValueError(
                    f"circulation facility {item_id} is rewarded before its materials at stage {stage_index}"
                )
            if item_id // 100 == 250:
                mine_circulation_ids.extend([item_id] * count)

    if mine_circulation_ids:
        raise ValueError(
            f"spirit stone vein must not appear in circulation rewards, got {mine_circulation_ids}"
        )

    breakthrough_reward_ids = []
    for cultivation_level, cultivation_stage in enumerate(stages[:-1], 1):
        reward_id = parse_int(cultivation_stage.get("breakthrough_reward_id"))
        if reward_id is None or str(reward_id) not in rewards:
            raise ValueError(
                f"cultivation level {cultivation_level} is missing a breakthrough reward"
            )
        reward_items = rewards[str(reward_id)].get("items", [])
        mine_count = sum(
            parse_int(item.get("count", 1)) or 0
            for item in reward_items
            if parse_int(item.get("id")) == SPIRIT_STONE_MINE_ID
        )
        if mine_count != 1:
            raise ValueError(
                f"cultivation level {cultivation_level} breakthrough must grant exactly one "
                f"{SPIRIT_STONE_MINE_ID}, got {mine_count}"
            )
        breakthrough_reward_ids.append(reward_id)
        for item in reward_items:
            item_id = parse_int(item.get("id"))
            if item_id in launcher_outputs or item_id in crafting_by_id:
                add_available_facility(item_id)

    # A launcher may intentionally be awarded before its first dependent table
    # (for example, the new spirit-wood line).  Check eventual usability after
    # the full staged schedule has been assembled instead of requiring a table
    # in the same circulation stage.
    unusable_launcher_ids = sorted(
        item_id
        for item_id in available_launcher_ids
        if item_id // 100 != SPIRIT_STONE_MINE_FAMILY
        and not launcher_has_usable_table(item_id, available_materials, available_table_ids)
    )
    if unusable_launcher_ids:
        raise ValueError(
            "circulation rewards leave launchers without a usable crafting table: "
            f"{unusable_launcher_ids}"
        )

    for reward_id in sorted(set(breakthrough_reward_ids)):
        for item in rewards.get(str(reward_id), {}).get("items", []):
            if item["count"] != 1:
                raise ValueError(f"breakthrough facility reward {reward_id} count must be 1")
            if not table_is_usable(item["id"], available_materials):
                raise ValueError(
                    f"breakthrough crafting table {item['id']} has no available materials"
                )


validate_facility_rewards()
save_json("home_meridians.json", {"stages": home_stages})

print(f"\nDone! JSON files written to {OUT}")
